# this file contain the BarCodeEAN13 class which can generate a code128 or EAN-13 bar code
from openpyxl import load_workbook


class BarCodeEAN13:
    def __init__(self, information):  # add the type
        self.information = information
        self.first_number = [["A", "A", "A", "A", "A", "A"],
                             ["A", "A", "B", "A", "B", "B"],
                             ["A", "A", "B", "B", "A", "B"],
                             ["A", "A", "B", "B", "B", "A"],
                             ["A", "B", "A", "A", "B", "B"],
                             ["A", "B", "B", "A", "A", "B"],
                             ["A", "B", "B", "B", "A", "A"],
                             ["A", "B", "A", "B", "A", "B"],
                             ["A", "B", "A", "B", "B", "A"],
                             ["A", "B", "B", "A", "B", "A"]]  # see this for more information :
        # https://grandzebu.net/informatique/codbar-en/ean13.htm

        self.core_list = ["101"]  # this list is the bar code : 1 stands for a dark line and 0 for a white one
        wb = load_workbook("Exels files/Correspondence.xlsx")
        ws = wb.active

        for i in range(1, 7):
            self.core_list.append(ws["{}{}".format(chr(ord(self.first_number[int(self.information[0])][i - 1]) + 1), int(self.information[i]) + 2)].value)

        self.core_list.append("01010")
        for i in range(7, 13):
            self.core_list.append(ws["{}{}".format("D", int(self.information[i]) + 2)].value)
        self.core_list.append("101")

    def check_digit(self):
        sum_even_nb = 0
        sum_odd_nb = 0

        for i in range(0, len(self.information) - 1, 2):
            if i != len(self.information):
                sum_odd_nb += int(self.information[i])

        for j in range(1, len(self.information), 2):
            sum_even_nb += 3*int(self.information[j])

        total = sum_odd_nb + sum_even_nb
        total = total % 10
        if total != 0:
            total = 10 - total

        if total == int(self.information[-1]):
            return True
        else:

            return False

