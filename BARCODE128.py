# this file contain the class BarCode128 which is a type a bar code able to stock letters
# this class use the 128B version only
from openpyxl import load_workbook


class BarCode128:
    def __init__(self, information):
        self.information = information
        self.core_list = ["11010010000"]  # this is the list of binary information to give to the draw function and we
        # add the START value
        wb = load_workbook("Exels files/Code128_Codes.xlsx")
        ws = wb["Main worksheet"]
        control_key = 104  # this is made to make sure there is no error in the information read

        for i in range(len(self.information)):

            for j in range(2, 99):

                if self.information[i] == ws["A{}".format(j)].value:

                    self.core_list.append(str(ws["B{}".format(j)].value))
                    control_key += ((j - 2) * (i + 1))

        control_key = control_key % 103

        if control_key != 1 and control_key != 0:
            control_key = ws["B{}".format(control_key + 2)].value  # we add 2 because the file Code128_codes is shifted
            # by 2 characters

        else:
            if control_key == 1:
                control_key = "11001101100"

            elif control_key == 0:
                control_key = "11011001100"

        self.core_list.append(control_key)
        self.core_list.append("1100011101011")  # we add the STOP


